import pygame
import sys
import time
import random
import re
import math
from collections import deque
from openpyxl import load_workbook
from routing import Label, Ship, pareto_optimal_path, reconstruct_path, prune_path

# Initialize Pygame
pygame.init()

# Load arctic map image first to get its dimensions for window size
try:
    temp_img = pygame.image.load("visuals/islands.png")
    img_width, img_height = temp_img.get_size()
    width, height = img_width, img_height
except pygame.error as e:
    print(f"Couldn't load islands image for sizing: {e}")
    width, height = 800, 600  # Fallback to default size

# Set up the display
screen = pygame.display.set_mode((width, height))
pygame.display.set_caption("Ship Selection")

# Define colors
WHITE = (255, 255, 255)
BLACK = (0, 0, 0)
BLUE = (100, 150, 255)
DARK_BLUE = (50, 100, 200)
GRAY = (200, 200, 200)
LIGHT_GRAY = (240, 240, 240)
GREEN = (50, 200, 50)
DARK_GREEN = (30, 150, 30)
RED = (255, 0, 0)
ORANGE = (255, 165, 0)

# Global grid cache to avoid redundant analysis
_cached_grid = None

# Load background image
try:
    background_img = pygame.image.load("visuals/arctic-ice-ridges.jpg")
    background_img = pygame.transform.scale(background_img, (width, height))
except pygame.error as e:
    print(f"Couldn't load background image: {e}")
    background_img = None

# Load ship placeholder image
try:
    ship_placeholder_img = pygame.image.load("visuals/salacola.png")
except pygame.error as e:
    print(f"Couldn't load ship placeholder image: {e}")
    ship_placeholder_img = None

# Load arctic map image for page 2 (already loaded for sizing, reload for use)
arctic_map_img = None
try:
    arctic_map_img = pygame.image.load("visuals/islands.png")
    # No scaling needed - window is already the same size as the image
except pygame.error as e:
    print(f"Couldn't load arctic map image: {e}")
    arctic_map_img = None

# Grid spacing constant
GRID_SPACING = 25

# GridCell class to store base properties of a grid square
class GridCell:
    def __init__(self, risk=0.0, time=1.0, fuel=1.0, weather=1.0, is_clickable=False):
        self.risk = risk
        self.time = time
        self.fuel = fuel
        self.weather = weather
        self.is_clickable = is_clickable

# Function to check if a point is on blue (water) surface
def is_blue_surface(x, y):
    """Check if the pixel at (x, y) is blue (water)"""
    if arctic_map_img is None:
        return False
    
    # Make sure coordinates are within image bounds
    if x < 0 or y < 0 or x >= width or y >= height:
        return False
    
    try:
        # Get pixel color at the click position
        pixel_color = arctic_map_img.get_at((x, y))
        r, g, b, a = pixel_color
        
        # Check if it's blue (blue channel is dominant)
        # Stricter margin to avoid gray (where r, g, b are almost equal)
        is_blue = b > r + 15 and b > g + 15 and b > 100
        
        return is_blue
    except:
        return False

def analyze_cell_from_image(cell_x, cell_y, grid_spacing, pixel_array=None):
    """
    Analyze image pixels in a grid cell to determine its properties.
    Returns a GridCell instance.
    """
    if arctic_map_img is None:
        return GridCell(risk=1.0, time=1.0, fuel=1.0, weather=1.0)
    
    # Sample pixels within the cell (5x5 grid)
    blue_pixels = 0
    total_samples = 0
    total_brightness = 0
    
    # Use PixelArray if provided for much faster access
    use_pixel_array = pixel_array is not None
    
    for i in range(5):
        for j in range(5):
            px = cell_x + (i + 1) * (grid_spacing // 6)
            py = cell_y + (j + 1) * (grid_spacing // 6)
            
            if 0 <= px < width and 0 <= py < height:
                try:
                    if use_pixel_array:
                        raw_color = pixel_array[px, py]
                        # Use surface.unmap_rgb for safe cross-platform RGB extraction
                        r, g, b = arctic_map_img.unmap_rgb(raw_color)[:3]
                    else:
                        color = arctic_map_img.get_at((px, py))
                        r, g, b, _ = color
                    
                    # Blue detection: b > r + 15 and b > g + 15 and b > 100
                    if b > r + 15 and b > g + 15 and b > 100:
                        blue_pixels += 1
                    
                    total_brightness += (r + g + b) / 3
                    total_samples += 1
                except:
                    continue
    
    if total_samples == 0:
        return GridCell(risk=5.0, time=3.0, fuel=2.0, weather=1.0, is_clickable=False)
    
    water_ratio = blue_pixels / total_samples
    ice_ratio = 1.0 - water_ratio
    avg_brightness = total_brightness / total_samples
    
    # Risk: base risk from ice, plus brightness bonus for thick ice, plus significant independent randomness
    risk = ice_ratio * 5.0 + (avg_brightness / 255.0) * 2.0 + random.uniform(0, 10.0)
    
    # Time multiplier: base ice delay plus significant independent randomness
    time_mult = 1.0 + ice_ratio * 5.0 + random.uniform(0, 10.0)
    
    # Fuel multiplier: base ice drag plus significant independent randomness
    fuel_mult = 1.0 + ice_ratio * 3.0 + random.uniform(0, 7.0)
    
    # Weather: simulated as a combination of ice and randomness
    weather = 1.0 + ice_ratio * 2.0 + random.uniform(0, 5.0)
    
    # Determine if clickable: show grid if ANY sampled point is water
    # This helps catch narrow rivers and small bodies of water
    is_clickable = blue_pixels >= 1
    
    return GridCell(risk=risk, time=time_mult, fuel=fuel_mult, weather=weather, is_clickable=is_clickable)

def init_grid_cells(width, height, grid_spacing):
    """
    Initialize grid cells and perform reachability check from top-left.
    Uses PixelArray for performance and caches the result.
    """
    global _cached_grid
    if _cached_grid is not None:
        return _cached_grid

    grid_cols = (width + grid_spacing - 1) // grid_spacing
    grid_rows = (height + grid_spacing - 1) // grid_spacing
    grid = []
    
    # Use PixelArray to lock the surface for high-speed access
    pixel_array = pygame.PixelArray(arctic_map_img)
    
    try:
        # 1. Initial analysis
        for row in range(grid_rows):
            grid_row = []
            for col in range(grid_cols):
                cell_x = col * grid_spacing
                cell_y = row * grid_spacing
                cell_data = analyze_cell_from_image(cell_x, cell_y, grid_spacing, pixel_array)
                grid_row.append(cell_data)
            grid.append(grid_row)
    finally:
        # Unlock the surface
        pixel_array.close()
        
    # 2. Find starting point (first clickable cell from top-left)
    start_node = None
    for r in range(grid_rows):
        for c in range(grid_cols):
            if grid[r][c].is_clickable:
                start_node = (r, c)
                break
        if start_node:
            break
            
    if not start_node:
        _cached_grid = grid
        return grid

    # 3. BFS Reachability Check
    reachable = set()
    queue = deque([start_node])
    reachable.add(start_node)
    
    while queue:
        r, c = queue.popleft()
        # Neighbors: Up, Down, Left, Right
        for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
            nr, nc = r + dr, c + dc
            if 0 <= nr < grid_rows and 0 <= nc < grid_cols:
                if grid[nr][nc].is_clickable and (nr, nc) not in reachable:
                    reachable.add((nr, nc))
                    queue.append((nr, nc))
                    
    # 4. Filter non-reachable cells
    for r in range(grid_rows):
        for c in range(grid_cols):
            if (r, c) not in reachable:
                grid[r][c].is_clickable = False
                
    _cached_grid = grid
    return grid

def get_spline_points(path, num_segments=15):
    """
    Generate Catmull-Rom spline points for a given path to make it look smooth.
    """
    if len(path) < 2:
        return path
    
    # Duplicate start and end points to handle Catmull-Rom boundaries
    points = [path[0]] + path + [path[-1]]
    
    spline_path = []
    for i in range(1, len(points) - 2):
        p0 = points[i-1]
        p1 = points[i]
        p2 = points[i+1]
        p3 = points[i+2]
        
        for t_int in range(num_segments):
            t = t_int / num_segments
            t2 = t * t
            t3 = t2 * t
            
            # Catmull-Rom interpolation formula
            x = 0.5 * (
                (2 * p1[0]) +
                (-p0[0] + p2[0]) * t +
                (2 * p0[0] - 5 * p1[0] + 4 * p2[0] - p3[0]) * t2 +
                (-p0[0] + 3 * p1[0] - 3 * p2[0] + p3[0]) * t3
            )
            y = 0.5 * (
                (2 * p1[1]) +
                (-p0[1] + p2[1]) * t +
                (2 * p0[1] - 5 * p1[1] + 4 * p2[1] - p3[1]) * t2 +
                (-p0[1] + 3 * p1[1] - 3 * p2[1] + p3[1]) * t3
            )
            spline_path.append((x, y))
            
    # Add the final point
    spline_path.append(path[-1])
    return spline_path

def get_final_path_points(label, grid_cells, point_a, point_b, goal_node, grid_spacing):
    """
    Process a Pareto label into a final list of screen coordinates.
    Handles reconstruction, pruning, coordinate conversion, and spline validation.
    """
    if not label:
        return None
        
    # 1. Reconstruct initial grid path: list of (r, c)
    raw_path_indices = reconstruct_path(label)
    raw_path_indices.append(goal_node)
    
    # 2. Prune the path to get direct any-angle lines (String Pulling)
    pruned_indices = prune_path(grid_cells, raw_path_indices)
    
    # 3. Convert grid indices to screen coordinates
    waypoint_coords = []
    # Start directly from point_a (no hook to cell center)
    waypoint_coords.append(point_a)
    
    # Convert intermediate waypoints to centers
    for r, c in pruned_indices:
        px = c * grid_spacing + grid_spacing // 2
        py = r * grid_spacing + grid_spacing // 2
        
        # Avoid adding point if it's too close to point_a or point_b
        dist_a = math.sqrt((px - point_a[0])**2 + (py - point_a[1])**2)
        dist_b = math.sqrt((px - point_b[0])**2 + (py - point_b[1])**2)
        
        if dist_a > grid_spacing // 2 and dist_b > grid_spacing // 2:
            waypoint_coords.append((px, py))
    
    # End directly at point_b
    waypoint_coords.append(point_b)
    
    # 4. Apply Catmull-Rom Spline smoothing for a natural curve
    raw_spline_path = get_spline_points(waypoint_coords, num_segments=15)
    
    # 5. Final Validation: Ensure smoothed points don't drift onto land
    # Optimization: Sample only 1 out of every 3 spline points for validation
    is_path_valid = True
    for i in range(0, len(raw_spline_path), 3):
        px, py = raw_spline_path[i]
        if not is_blue_surface(int(px), int(py)):
            is_path_valid = False
            break
    
    if is_path_valid and not is_blue_surface(int(raw_spline_path[-1][0]), int(raw_spline_path[-1][1])):
        is_path_valid = False
    
    if is_path_valid:
        return raw_spline_path
    else:
        # Fallback to pruned waypoints if spline drifts
        return waypoint_coords

# Load ships data from Excel
def load_ships_data():
    ships = []
    try:
        wb = load_workbook('Ships.xlsx')
        ws = wb.active
        
        # Read header row - columns are at indices 0, 2, 4, 6, 7, 9
        header_row = list(ws[1])
        headers = {
            0: header_row[0].value.strip() if header_row[0].value else None,  # Ship type
            2: header_row[2].value.strip() if header_row[2].value else None,  # Ship name
            4: header_row[4].value.strip() if header_row[4].value else None,  # Fuel Consumption
            6: header_row[6].value.strip() if header_row[6].value else None,  # Speed
            7: header_row[7].value.strip() if header_row[7].value else None,  # Durability
            9: header_row[9].value.strip() if header_row[9].value else None   # Durability Rating
        }
        
        # Read ship data dynamically
        for row_idx in range(2, ws.max_row + 1):  # Read all rows starting from row 2
            row = list(ws[row_idx])
            if not row[0].value:  # Stop if ship type is missing (empty row)
                continue
                
            ship_data = {}
            
            # Map data to headers based on column indices
            if row[0].value:  # Ship type
                ship_data[headers[0]] = str(row[0].value).strip()
            if row[2].value:  # Ship name
                ship_data[headers[2]] = str(row[2].value).strip()
            if row[4].value:  # Fuel Consumption
                ship_data[headers[4]] = str(row[4].value).strip()
            if row[6].value:  # Speed
                ship_data[headers[6]] = str(row[6].value).strip()
            if row[7].value:  # Durability
                ship_data[headers[7]] = str(row[7].value).strip()
            if row[9].value is not None:  # Durability Rating (can be a number)
                ship_data[headers[9]] = str(row[9].value).strip()
            
            if ship_data:  # Only add if we have data
                # Convert to a Ship object for the routing algorithm if all required fields are present
                try:
                    # Clean the data (remove units if present, etc.)
                    def clean_val(val, default=0.0):
                        if not val: return default
                        # Extract first number found in string
                        match = re.search(r"[-+]?\d*\.\d+|\d+", str(val))
                        return float(match.group()) if match else default

                    # Store the original data for display and the Ship object for logic
                    ship_data['obj'] = Ship(
                        base_speed=clean_val(ship_data.get('Speed', '0'), default=1.0),
                        base_fuel_rate=clean_val(ship_data.get('Fuel Consumption', '0'), default=1.0),
                        durability=clean_val(ship_data.get('Durability', '0'), default=1.0)
                    )
                except Exception as e:
                    print(f"Warning: Could not create Ship object for {ship_data.get('Ship name', 'unknown')}: {e}")
                
                ships.append(ship_data)
    except Exception as e:
        print(f"Error loading ships data: {e}")
        import traceback
        traceback.print_exc()
    
    return ships

# Load ships
ships = load_ships_data()
selected_ship = None
current_page = 1  # 1 = selection page, 2 = next page
confirmed_ship = None
selected_ship_type = None  # Store the selected ship type
page2_start_time = None  # Track when page 2 was entered
point_a = None  # Store point A position
point_b = None  # Store point B position
points_confirmed = False  # Track if points A and B have been confirmed
toggle_on = False  # Toggle button state
grid_cells = None  # 2D grid of GridCell instances for each grid square
available_paths = {'time': None, 'fuel': None, 'risk': None}  # Dictionary to store optimized paths
selected_path_type = 'time'  # Currently selected path type
selected_grid_layer = 'none'  # Grid visualization layer: 'none', 'risk', 'time', 'fuel'

def get_value_color(value, min_val, max_val):
    """
    Calculate a color from Green (min) to Red (max).
    Returns (r, g, b, alpha)
    """
    if max_val == min_val:
        return (0, 255, 0, 128)
    
    # Normalize value between 0 and 1
    t = max(0, min(1, (value - min_val) / (max_val - min_val)))
    
    # Interpolate between Green (0, 255, 0) and Red (255, 0, 0)
    r = int(255 * t)
    g = int(255 * (1 - t))
    b = 0
    return (r, g, b, 128)

# Font setup
font_large = pygame.font.Font(None, 36)
font_medium = pygame.font.Font(None, 24)
font_small = pygame.font.Font(None, 20)

# Helper function to draw text with word wrapping
def draw_text_wrapped(surface, text, font, color, rect, aa=True):
    words = text.split(' ')
    lines = []
    current_line = []
    
    for word in words:
        test_line = ' '.join(current_line + [word])
        test_width, _ = font.size(test_line)
        if test_width <= rect.width:
            current_line.append(word)
        else:
            if current_line:
                lines.append(' '.join(current_line))
            current_line = [word]
    if current_line:
        lines.append(' '.join(current_line))
    
    y_offset = 0
    for line in lines:
        if y_offset + font.get_height() > rect.height:
            break
        text_surface = font.render(line, aa, color)
        surface.blit(text_surface, (rect.x, rect.y + y_offset))
        y_offset += font.get_height() + 2

# Helper function to format ship description
def format_ship_description(ship):
    description = []
    
    # Define the order and labels for each field
    fields = [
        ('Ship type', 'Ship type'),
        ('Ship name', 'Ship name'),
        ('Fuel Consumption         ', 'Fuel Consumption'),
        ('Speed', 'Speed'),
        ('Durability', 'Durability'),
        ('Durability Rating', 'Durability Rating')
    ]
    
    for excel_key, display_label in fields:
        value = ship.get(excel_key, '')
        if value and value != 'None' and str(value).strip():
            description.append(f"{display_label}:")
            description.append(f"  {value}")
            description.append("")  # Empty line for spacing
    
    return "\n".join(description)

# Helper function to draw ship description with proper formatting
def draw_ship_description(surface, ship, font_label, font_value, color, rect):
    y_offset = 0
    line_height = font_label.get_height() + 4
    
    # Define the order and labels for each field
    fields = [
        ('Ship type', 'Ship type'),
        ('Ship name', 'Ship name'),
        ('Fuel Consumption         ', 'Fuel Consumption'),
        ('Speed', 'Speed'),
        ('Durability', 'Durability'),
        ('Durability Rating', 'Durability Rating')
    ]
    
    for excel_key, display_label in fields:
        value = ship.get(excel_key, '')
        if value and value != 'None' and str(value).strip():
            # Draw label
            if y_offset + line_height > rect.height:
                break
            label_text = font_label.render(f"{display_label}:", True, color)
            surface.blit(label_text, (rect.x, rect.y + y_offset))
            y_offset += line_height
            
            # Draw value (with word wrapping if needed)
            if y_offset + line_height > rect.height:
                break
            words = str(value).split(' ')
            current_line = []
            for word in words:
                test_line = ' '.join(current_line + [word])
                test_width, _ = font_value.size(test_line)
                if test_width <= rect.width - 20:  # Leave some margin
                    current_line.append(word)
                else:
                    if current_line:
                        value_text = font_value.render('  ' + ' '.join(current_line), True, color)
                        surface.blit(value_text, (rect.x, rect.y + y_offset))
                        y_offset += line_height
                        if y_offset + line_height > rect.height:
                            break
                    current_line = [word]
            
            if current_line:
                value_text = font_value.render('  ' + ' '.join(current_line), True, color)
                surface.blit(value_text, (rect.x, rect.y + y_offset))
                y_offset += line_height
            
            # Add spacing between fields
            y_offset += 8

# Main game loop
running = True
while running:
    # Handle events
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
        elif event.type == pygame.KEYDOWN:
            if event.key == pygame.K_r:
                # Reset all state variables to restart the program
                selected_ship = None
                current_page = 1
                confirmed_ship = None
                selected_ship_type = None
                page2_start_time = None
                point_a = None
                point_b = None
                points_confirmed = False
                toggle_on = False
                selected_grid_layer = 'none'
                grid_cells = None
                available_paths = {'time': None, 'fuel': None, 'risk': None}
                selected_path_type = 'time'
        elif event.type == pygame.MOUSEBUTTONDOWN:
            mouse_x, mouse_y = event.pos
            
            if current_page == 1:  # Selection page
                if event.button == 1:  # Left click only on page 1
                    # Check if clicking on confirmation button (in description panel)
                    if selected_ship:
                        panel_x = 300
                        panel_y = 60
                        panel_width = width - panel_x - 50
                        panel_height = height - panel_y - 50
                        button_width = 180
                        button_height = 40
                        confirm_button_x = panel_x + panel_width - button_width - 15
                        confirm_button_y = panel_y + panel_height - button_height - 15
                        confirm_button_rect = pygame.Rect(confirm_button_x, confirm_button_y, button_width, button_height)
                        
                        if confirm_button_rect.collidepoint(mouse_x, mouse_y):
                            confirmed_ship = selected_ship
                            selected_ship_type = selected_ship.get('Ship type', 'Unknown') if selected_ship else None
                            current_page = 2
                            page2_start_time = time.time()  # Record when page 2 starts
                            # Initialize grid cells when entering page 2
                            grid_cells = init_grid_cells(width, height, GRID_SPACING)
                            continue
                    
                    # Check if clicking on a ship button
                    button_width = 250
                    button_height = 50
                    button_start_y = 60
                    button_spacing = 60
                    
                    for i, ship in enumerate(ships):
                        button_x = 50
                        button_y = button_start_y + i * button_spacing
                        button_rect = pygame.Rect(button_x, button_y, button_width, button_height)
                        
                        if button_rect.collidepoint(mouse_x, mouse_y):
                            selected_ship = ship
                            break
            
            elif current_page == 2:  # Page 2: Map clicking
                if event.button == 1:  # Left click
                    # 1. UI Buttons First
                    # Toggle Grid button
                    toggle_button_width = 130
                    toggle_button_height = 30
                    toggle_button_x = width - toggle_button_width - 10
                    toggle_button_y = 10
                    toggle_button_rect = pygame.Rect(toggle_button_x, toggle_button_y, toggle_button_width, toggle_button_height)
                    
                    # Grid Layer buttons
                    layer_btn_w = 130
                    layer_btn_h = 30
                    layer_btn_x = width - layer_btn_w - 10
                    time_rect = pygame.Rect(layer_btn_x, 50, layer_btn_w, layer_btn_h)
                    fuel_rect = pygame.Rect(layer_btn_x, 90, layer_btn_w, layer_btn_h)
                    risk_rect = pygame.Rect(layer_btn_x, 130, layer_btn_w, layer_btn_h)
                    
                    # Route selection buttons (bottom left)
                    route_btn_w = 200
                    route_btn_h = 45
                    route_btn_x = 10
                    route_btn_y_start = height - (route_btn_h + 10) * 3 - 10
                    fastest_rect = pygame.Rect(route_btn_x, route_btn_y_start, route_btn_w, route_btn_h)
                    eco_rect = pygame.Rect(route_btn_x, route_btn_y_start + route_btn_h + 10, route_btn_w, route_btn_h)
                    safest_rect = pygame.Rect(route_btn_x, route_btn_y_start + (route_btn_h + 10) * 2, route_btn_w, route_btn_h)
                    
                    # Map Confirm button
                    confirm_btn_w = 150
                    confirm_btn_h = 40
                    confirm_btn_x = 10
                    confirm_btn_y = height - confirm_btn_h - 10
                    map_confirm_rect = pygame.Rect(confirm_btn_x, confirm_btn_y, confirm_btn_w, confirm_btn_h)

                    # Handle Clicks
                    if toggle_button_rect.collidepoint(mouse_x, mouse_y):
                        toggle_on = not toggle_on
                        if toggle_on and grid_cells is None:
                            grid_cells = init_grid_cells(width, height, GRID_SPACING)
                        if not toggle_on:
                            selected_grid_layer = 'none'
                    elif toggle_on and time_rect.collidepoint(mouse_x, mouse_y):
                        selected_grid_layer = 'time' if selected_grid_layer != 'time' else 'none'
                    elif toggle_on and fuel_rect.collidepoint(mouse_x, mouse_y):
                        selected_grid_layer = 'fuel' if selected_grid_layer != 'fuel' else 'none'
                    elif toggle_on and risk_rect.collidepoint(mouse_x, mouse_y):
                        selected_grid_layer = 'risk' if selected_grid_layer != 'risk' else 'none'
                    elif points_confirmed and any(available_paths.values()):
                        if fastest_rect.collidepoint(mouse_x, mouse_y):
                            selected_path_type = 'time'
                        elif eco_rect.collidepoint(mouse_x, mouse_y):
                            selected_path_type = 'fuel'
                        elif safest_rect.collidepoint(mouse_x, mouse_y):
                            selected_path_type = 'risk'
                    elif not points_confirmed:
                        if point_a and point_b and map_confirm_rect.collidepoint(mouse_x, mouse_y):
                            # Visual feedback for processing
                            overlay = pygame.Surface((width, height))
                            overlay.set_alpha(128)
                            overlay.fill(BLACK)
                            screen.blit(overlay, (0, 0))
                            proc_text = font_large.render("Processing Path...", True, WHITE)
                            proc_rect = proc_text.get_rect(center=(width // 2, height // 2))
                            screen.blit(proc_text, proc_rect)
                            pygame.display.flip()
                            
                            points_confirmed = True
                            if grid_cells:
                                grid_spacing = GRID_SPACING
                                start_node = (point_a[1] // grid_spacing, point_a[0] // grid_spacing)
                                goal_node = (point_b[1] // grid_spacing, point_b[0] // grid_spacing)
                                start_node = (max(0, min(len(grid_cells)-1, start_node[0])), max(0, min(len(grid_cells[0])-1, start_node[1])))
                                goal_node = (max(0, min(len(grid_cells)-1, goal_node[0])), max(0, min(len(grid_cells[0])-1, goal_node[1])))
                                pareto_labels = pareto_optimal_path(grid_cells, start_node, goal_node, confirmed_ship['obj'])
                                if pareto_labels:
                                    time_label = min(pareto_labels, key=lambda l: l.time)
                                    fuel_label = min(pareto_labels, key=lambda l: l.fuel)
                                    risk_label = min(pareto_labels, key=lambda l: l.risk)
                                    available_paths['time'] = {'points': get_final_path_points(time_label, grid_cells, point_a, point_b, goal_node, grid_spacing), 'time': time_label.time, 'fuel': time_label.fuel, 'risk': time_label.risk}
                                    available_paths['fuel'] = {'points': get_final_path_points(fuel_label, grid_cells, point_a, point_b, goal_node, grid_spacing), 'time': fuel_label.time, 'fuel': fuel_label.fuel, 'risk': fuel_label.risk}
                                    available_paths['risk'] = {'points': get_final_path_points(risk_label, grid_cells, point_a, point_b, goal_node, grid_spacing), 'time': risk_label.time, 'fuel': risk_label.fuel, 'risk': risk_label.risk}
                                    selected_path_type = 'time'
                                else:
                                    available_paths = {'time': None, 'fuel': None, 'risk': None}
                        elif is_blue_surface(mouse_x, mouse_y):
                            point_a = (mouse_x, mouse_y)
                            available_paths = {'time': None, 'fuel': None, 'risk': None}
                elif event.button == 3:  # Right click - point B
                    if not points_confirmed:
                        if is_blue_surface(mouse_x, mouse_y):
                            point_b = (mouse_x, mouse_y)
                            available_paths = {'time': None, 'fuel': None, 'risk': None}

    # Draw background
    if current_page == 1:
        # Page 1: Show arctic background
        if background_img:
            screen.blit(background_img, (0, 0))
        else:
            screen.fill(WHITE)
    elif current_page == 2:
        # Page 2: Arctic map background
        if arctic_map_img:
            screen.blit(arctic_map_img, (0, 0))
        else:
            screen.fill(WHITE)

    # Page 1: Ship Selection
    if current_page == 1:
        # Draw title
        title_text = font_large.render("Select a Ship", True, WHITE)
        screen.blit(title_text, (50, 10))

        # Draw ship selection buttons
        button_width = 250
        button_height = 50
        button_start_y = 60
        button_spacing = 60
        
        for i, ship in enumerate(ships):
            button_x = 50
            button_y = button_start_y + i * button_spacing
            button_rect = pygame.Rect(button_x, button_y, button_width, button_height)
            
            # Highlight selected ship
            if selected_ship == ship:
                pygame.draw.rect(screen, DARK_BLUE, button_rect)
            else:
                pygame.draw.rect(screen, BLUE, button_rect)
            
            # Draw button border
            pygame.draw.rect(screen, BLACK, button_rect, 2)
            
            # Draw ship name (full name, with word wrapping if needed)
            ship_name = ship.get('Ship name', f'Ship {i+1}')
            # Try to fit the full name, use smaller font if needed
            button_text = font_medium.render(ship_name, True, WHITE)
            text_width, text_height = button_text.get_size()
            
            # If text is too wide, use smaller font
            if text_width > button_width - 20:
                button_text = font_small.render(ship_name, True, WHITE)
                text_width, text_height = button_text.get_size()
                # If still too wide, wrap to multiple lines
                if text_width > button_width - 20:
                    words = ship_name.split(' ')
                    lines = []
                    current_line = []
                    for word in words:
                        test_line = ' '.join(current_line + [word])
                        test_text = font_small.render(test_line, True, WHITE)
                        if test_text.get_width() <= button_width - 20:
                            current_line.append(word)
                        else:
                            if current_line:
                                lines.append(' '.join(current_line))
                            current_line = [word]
                    if current_line:
                        lines.append(' '.join(current_line))
                    
                    # Draw multiple lines
                    line_height = font_small.get_height() + 2
                    total_height = len(lines) * line_height
                    start_y = button_rect.centery - total_height // 2
                    for line in lines:
                        line_text = font_small.render(line, True, WHITE)
                        line_rect = line_text.get_rect(centerx=button_rect.centerx, y=start_y)
                        screen.blit(line_text, line_rect)
                        start_y += line_height
                else:
                    text_rect = button_text.get_rect(center=button_rect.center)
                    screen.blit(button_text, text_rect)
            else:
                text_rect = button_text.get_rect(center=button_rect.center)
                screen.blit(button_text, text_rect)

        # Draw description panel
        if selected_ship:
            panel_x = 300
            panel_y = 60
            panel_width = width - panel_x - 50
            panel_height = height - panel_y - 50
            
            # Draw semi-transparent panel background
            panel_surface = pygame.Surface((panel_width, panel_height))
            panel_surface.set_alpha(230)
            panel_surface.fill(LIGHT_GRAY)
            screen.blit(panel_surface, (panel_x, panel_y))
            
            # Draw panel border
            panel_rect = pygame.Rect(panel_x, panel_y, panel_width, panel_height)
            pygame.draw.rect(screen, BLACK, panel_rect, 2)
            
            # Draw description title
            title_text = font_large.render("Ship Details", True, BLACK)
            screen.blit(title_text, (panel_x + 10, panel_y + 10))
            
            # Draw ship image (placeholder)
            image_size = 150
            image_x = panel_x + 10
            image_y = panel_y + 50
            if ship_placeholder_img:
                # Scale image to fit
                img_width, img_height = ship_placeholder_img.get_size()
                scale = min(image_size / img_width, image_size / img_height)
                scaled_width = int(img_width * scale)
                scaled_height = int(img_height * scale)
                scaled_img = pygame.transform.scale(ship_placeholder_img, (scaled_width, scaled_height))
                screen.blit(scaled_img, (image_x, image_y))
            
            # Draw ship description (leave space for image and button at bottom)
            # Position text to the right of the image, or below if image is wide
            desc_x = panel_x + 10 + image_size + 10
            desc_y = panel_y + 50
            desc_width = panel_width - 20 - image_size - 10
            desc_height = panel_height - 70
            desc_rect = pygame.Rect(desc_x, desc_y, desc_width, desc_height)
            draw_ship_description(screen, selected_ship, font_medium, font_small, BLACK, desc_rect)

            # Draw green confirmation button in bottom right of description panel
            button_width = 180
            button_height = 40
            confirm_button_x = panel_x + panel_width - button_width - 15
            confirm_button_y = panel_y + panel_height - button_height - 15
            confirm_button_rect = pygame.Rect(confirm_button_x, confirm_button_y, button_width, button_height)
            
            # Draw button
            pygame.draw.rect(screen, GREEN, confirm_button_rect)
            pygame.draw.rect(screen, BLACK, confirm_button_rect, 2)
            
            # Draw button text
            confirm_text = font_medium.render("Confirm Selection", True, WHITE)
            text_rect = confirm_text.get_rect(center=confirm_button_rect.center)
            screen.blit(confirm_text, text_rect)

    # Page 2: Next Page
    elif current_page == 2:
        # Draw grid first (if toggle is on) so UI elements appear on top
        if toggle_on:
            grid_spacing = GRID_SPACING  # Grid cell size
            grid_color = (200, 200, 200, 128)  # Semi-transparent gray
            
            # Initialize grid cells if they don't exist
            if grid_cells is None:
                grid_cells = init_grid_cells(width, height, grid_spacing)
            
            # Draw horizontal lines (removed continuous lines)
            
            # Display risk, time, and fuel values in each cell
            if grid_cells is not None:
                font_cell = pygame.font.Font(None, 16)  # Font for cell values
                for row in range(len(grid_cells)):
                    for col in range(len(grid_cells[row])):
                        cell_data = grid_cells[row][col]
                        
                        # Only draw grid and numbers for clickable cells
                        if cell_data.is_clickable:
                            cell_x = col * grid_spacing
                            cell_y = row * grid_spacing
                            
                            # Draw cell boundary
                            cell_rect = pygame.Rect(cell_x, cell_y, grid_spacing, grid_spacing)
                            
                            # Fill cell if a layer is selected
                            if selected_grid_layer != 'none':
                                if selected_grid_layer == 'risk':
                                    color = get_value_color(cell_data.risk, 0.0, 15.0)
                                elif selected_grid_layer == 'time':
                                    color = get_value_color(cell_data.time, 1.0, 15.0)
                                elif selected_grid_layer == 'fuel':
                                    color = get_value_color(cell_data.fuel, 1.0, 10.0)
                                
                                # Draw filled rect with alpha using a temporary surface
                                s = pygame.Surface((grid_spacing, grid_spacing), pygame.SRCALPHA)
                                s.fill(color)
                                screen.blit(s, (cell_x, cell_y))
                                
                            pygame.draw.rect(screen, grid_color[:3], cell_rect, 1)
        
        # Draw the calculated path if it exists
        path_data = available_paths.get(selected_path_type)
        if path_data and path_data['points']:
            path_points = path_data['points']
            # Color map for the paths
            path_colors = {'time': GREEN, 'fuel': BLUE, 'risk': ORANGE}
            color = path_colors.get(selected_path_type, BLUE)
            
            # Draw a thick line connecting the path points
            if len(path_points) > 1:
                pygame.draw.lines(screen, color, False, path_points, 5)
        
        # Display ship name in top left (drawn on top of grid)
        if confirmed_ship:
            ship_name = confirmed_ship.get('Ship name', 'Unknown Ship')
            ship_name_text = font_large.render(ship_name, True, BLACK)
            screen.blit(ship_name_text, (10, 10))
        
        # Draw toggle button in top right (drawn on top of grid)
        toggle_button_width = 130
        toggle_button_height = 30
        toggle_button_x = width - toggle_button_width - 10
        toggle_button_y = 10
        toggle_button_rect = pygame.Rect(toggle_button_x, toggle_button_y, toggle_button_width, toggle_button_height)
        
        # Draw button with different color based on state
        if toggle_on:
            pygame.draw.rect(screen, GREEN, toggle_button_rect)
        else:
            pygame.draw.rect(screen, GRAY, toggle_button_rect)
        pygame.draw.rect(screen, BLACK, toggle_button_rect, 2)
        
        # Draw button text
        toggle_text = font_small.render("Show Grid", True, BLACK)
        text_rect = toggle_text.get_rect(center=toggle_button_rect.center)
        screen.blit(toggle_text, text_rect)
        
        # Draw layer selection buttons if grid is on
        if toggle_on:
            layer_btn_w = 130
            layer_btn_h = 30
            layer_btn_x = width - layer_btn_w - 10
            
            layers = [
                ("Time Layer", 'time', GREEN),
                ("Fuel Layer", 'fuel', BLUE),
                ("Risk Layer", 'risk', ORANGE)
            ]
            
            for i, (label, l_type, color) in enumerate(layers):
                btn_y = 50 + i * 40
                rect = pygame.Rect(layer_btn_x, btn_y, layer_btn_w, layer_btn_h)
                
                # Background
                bg_color = color if selected_grid_layer == l_type else LIGHT_GRAY
                pygame.draw.rect(screen, bg_color, rect)
                pygame.draw.rect(screen, BLACK, rect, 2)
                
                # Text
                txt_color = WHITE if selected_grid_layer == l_type else BLACK
                txt_surf = font_small.render(label, True, txt_color)
                txt_rect = txt_surf.get_rect(center=rect.center)
                screen.blit(txt_surf, txt_rect)
        
        # Display prompt for 3 seconds
        if page2_start_time is not None:
            elapsed_time = time.time() - page2_start_time
            if elapsed_time < 3:  # Show for 3 seconds
                # Create a semi-transparent background for the prompt
                prompt_bg = pygame.Surface((400, 60))
                prompt_bg.set_alpha(200)
                prompt_bg.fill(LIGHT_GRAY)
                prompt_x = width // 2 - 200
                prompt_y = height // 2 - 30
                screen.blit(prompt_bg, (prompt_x, prompt_y))
                
                # Draw prompt text
                prompt_text = font_medium.render("Select your points A and B", True, BLACK)
                text_rect = prompt_text.get_rect(center=(width // 2, height // 2))
                screen.blit(prompt_text, text_rect)
        
        # Draw point A if set
        if point_a:
            point_x, point_y = point_a
            # Draw red dot
            pygame.draw.circle(screen, RED, (point_x, point_y), 8)
            # Draw label "A" above the dot
            label_text = font_medium.render("A", True, BLACK)
            label_rect = label_text.get_rect(centerx=point_x, bottom=point_y - 12)
            screen.blit(label_text, label_rect)
        
        # Draw point B if set
        if point_b:
            point_x, point_y = point_b
            # Draw red dot
            pygame.draw.circle(screen, RED, (point_x, point_y), 8)
            # Draw label "B" above the dot
            label_text = font_medium.render("B", True, BLACK)
            label_rect = label_text.get_rect(centerx=point_x, bottom=point_y - 12)
            screen.blit(label_text, label_rect)
        
        # Draw confirm button in bottom left if both points are set and not yet confirmed
        if point_a and point_b and not points_confirmed:
            button_width = 150
            button_height = 40
            button_x = 10
            button_y = height - button_height - 10
            confirm_button_rect = pygame.Rect(button_x, button_y, button_width, button_height)
            
            # Draw button
            pygame.draw.rect(screen, GREEN, confirm_button_rect)
            pygame.draw.rect(screen, BLACK, confirm_button_rect, 2)
            
            # Draw button text
            confirm_text = font_medium.render("Confirm", True, WHITE)
            text_rect = confirm_text.get_rect(center=confirm_button_rect.center)
            screen.blit(confirm_text, text_rect)
            
        # Draw path selection buttons if points are confirmed
        if points_confirmed and any(available_paths.values()):
            button_width = 200
            button_height = 45
            button_x = 10
            button_y_start = height - (button_height + 10) * 3 - 10
            
            # Paths labels and types
            paths = [
                ("Fastest", 'time', GREEN),
                ("Eco-Friendly", 'fuel', BLUE),
                ("Safest", 'risk', ORANGE)
            ]
            
            for i, (label, p_type, color) in enumerate(paths):
                rect = pygame.Rect(button_x, button_y_start + i * (button_height + 10), button_width, button_height)
                
                # Highlight if selected
                if selected_path_type == p_type:
                    pygame.draw.rect(screen, color, rect)
                    text_color = WHITE
                else:
                    pygame.draw.rect(screen, LIGHT_GRAY, rect)
                    text_color = BLACK
                
                pygame.draw.rect(screen, BLACK, rect, 2)
                
                # Draw label and metrics
                path_data = available_paths.get(p_type)
                if path_data:
                    if p_type == 'time':
                        metric_text = f"{label}: {path_data['time']:.1f}h"
                    elif p_type == 'fuel':
                        metric_text = f"{label}: {path_data['fuel']:.1f}L"
                    else:
                        metric_text = f"{label}: {path_data['risk']:.1f}R"
                else:
                    metric_text = label
                    
                text_surf = font_medium.render(metric_text, True, text_color)
                text_rect = text_surf.get_rect(center=rect.center)
                screen.blit(text_surf, text_rect)

    # Update the display
    pygame.display.flip()

# Quit Pygame
pygame.quit()
sys.exit()
